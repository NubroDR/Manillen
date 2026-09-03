import csv
import itertools as it
import shutil
from collections import defaultdict
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


DATA_DIR = Path(__file__).resolve().parent / 'data'
RESERVE_FILL = PatternFill(fill_type='solid', fgColor='FFF0CF')
RESERVE_TEXT_COLOR = '806B1D'


def load_all_players(csv_filename=str(DATA_DIR / 'AllPlayers.csv')):
    """
    Load all players from a CSV file with one column (no header).
    Returns a list of player names.
    """
    players = []
    try:
        with open(csv_filename, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                if row and row[0].strip():  # Skip empty rows
                    players.append(row[0].strip())
    except FileNotFoundError:
        print(f"Warning: File '{csv_filename}' not found. Using empty player list.")
    
    return players

def create_player_pool(active_players):
    """
    Create player pool with reserve players if needed to fill tables of 4.
    Returns: (active_pool, number_of_tables, reserve_players_added)
    """
    number_of_active_players = len(active_players)
    number_of_tables = number_of_active_players // 4
    reserve_players = 0
    active_pool = active_players[:]
    
    # Add reserve players if needed
    if number_of_active_players % 4 != 0:
        number_of_tables += 1
        reserve_players = number_of_tables * 4 - number_of_active_players
        for reserve_player in range(reserve_players):
            active_pool.append(f"Reserve {reserve_player + 1}")
    
    print(f'Aantal tafels: {number_of_tables}')
    print(f'Aangevuld met {reserve_players} reservespelers')
    
    return active_pool, number_of_tables, reserve_players

def load_pair_counts(csv_filename=str(DATA_DIR / 'pairings.csv')):
    """Load existing pair counts from CSV."""
    pair_counts = defaultdict(int)
    try:
        with open(csv_filename, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader, None)  # Skip header
            for row in reader:
                if row:
                    pair = tuple(sorted([row[0], row[1]]))
                    pair_counts[pair] = int(row[2])
    except FileNotFoundError:
        pass
    return pair_counts


def save_pairings_history(config, play_date=None, history_filename=str(DATA_DIR / 'pairings_history.csv')):
    """Save a table pairing configuration to a history CSV file."""
    if play_date is None:
        play_date = date.today().isoformat()
    elif isinstance(play_date, date):
        play_date = play_date.isoformat()

    file_exists = False
    try:
        with open(history_filename, 'r', encoding='utf-8') as _:
            file_exists = True
    except FileNotFoundError:
        pass

    with open(history_filename, 'a', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(['Date', 'Table', 'Players'])
        for table_number, group in enumerate(config, start=1):
            writer.writerow([play_date, table_number, ' | '.join(group)])

    print(f"Saved pairing history to {history_filename} for date {play_date}")


SCOREBLAD_DATE_CELLS = ('C2', 'I2', 'O2', 'U2')
# Each name block is laid out as two vertical team columns:
# left-team-top, left-team-bottom, right-team-top, right-team-bottom
SCOREBLAD_NAME_BLOCKS = [
    ('B10', 'B19', 'E10', 'E19'),
    ('H10', 'H19', 'K10', 'K19'),
    ('N10', 'N19', 'Q10', 'Q19'),
    ('T10', 'T19', 'W10', 'W19'),
]


def _normalize_play_date(play_date):
    if play_date is None:
        return date.today()
    if isinstance(play_date, date):
        return play_date
    if isinstance(play_date, str):
        try:
            return date.fromisoformat(play_date)
        except ValueError:
            pass
        if len(play_date) == 8 and play_date.isdigit():
            return date.fromisoformat(
                f"{play_date[:4]}-{play_date[4:6]}-{play_date[6:]}"
            )
    raise ValueError('play_date must be a date or string in YYYY-MM-DD / YYYYMMDD format')


def _clear_scoreblad_sheet(ws):
    for cell in SCOREBLAD_DATE_CELLS:
        ws[cell].value = None
    for block in SCOREBLAD_NAME_BLOCKS:
        for cell in block:
            ws[cell].value = None


def _load_pairings_history(play_date=None, history_filename=str(DATA_DIR / 'pairings_history.csv')):
    """Load pairings for a given date from pairings_history.csv."""
    play_date = _normalize_play_date(play_date).isoformat()
    pairings_by_table = {}

    try:
        with open(history_filename, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if len(row) < 3:
                    continue
                row_date = row[0].strip()
                if row_date != play_date:
                    continue
                try:
                    table_number = int(row[1])
                except ValueError:
                    continue
                if table_number < 1 or table_number > 4:
                    raise ValueError(f'Invalid table number {table_number} in history')
                if table_number in pairings_by_table:
                    raise ValueError(f'Duplicate table entry for table {table_number} on {play_date}')
                players = [p.strip() for p in row[2].split('|') if p.strip()]
                if len(players) != 4:
                    raise ValueError(
                        f'Expected 4 players for table {table_number} on {play_date}, got {len(players)}'
                    )
                pairings_by_table[table_number] = players
    except FileNotFoundError:
        raise FileNotFoundError(f'History file not found: {history_filename}')

    if not pairings_by_table:
        raise ValueError(f'No pairings found for date {play_date} in {history_filename}')

    return [pairings_by_table[key] for key in sorted(pairings_by_table)]


def _get_reserve_substitutions(reserve1=None, reserve2=None, reserve3=None):
    return {
        'Reserve 1': reserve1.strip() if isinstance(reserve1, str) and reserve1.strip() else None,
        'Reserve 2': reserve2.strip() if isinstance(reserve2, str) and reserve2.strip() else None,
        'Reserve 3': reserve3.strip() if isinstance(reserve3, str) and reserve3.strip() else None,
    }


def fill_scorebladen_from_history(
    play_date=None,
    history_filename=str(DATA_DIR / 'pairings_history.csv'),
    template_filename=str(DATA_DIR / 'Scorebladen.xlsx'),
    output_dir=None,
    reserve1=None,
    reserve2=None,
    reserve3=None,
):
    """Fill scorebladen from a past pairing history date."""
    pairings = _load_pairings_history(play_date, history_filename)
    return fill_scorebladen(
        pairings,
        play_date=play_date,
        template_filename=template_filename,
        output_dir=output_dir,
        reserve1=reserve1,
        reserve2=reserve2,
        reserve3=reserve3,
    )


def fill_scorebladen(
    pairings,
    play_date=None,
    template_filename=str(DATA_DIR / 'Scorebladen.xlsx'),
    output_dir=None,
    reserve1=None,
    reserve2=None,
    reserve3=None,
):
    """Fill the Excel scorebladen template based on the given table pairings.

    pairings: list of 4-player groups, one group per table.
    play_date: date or string; written into the date cells of each used sheet.
    template_filename: Excel template to read.
    output_dir: optional directory where the filled copy is saved.
    reserve1/reserve2/reserve3: optional substitute names for placeholders Reserve 1/2/3.

    Returns the output file path as a string.
    """
    play_date = _normalize_play_date(play_date)
    template_path = Path(template_filename)
    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    if not isinstance(pairings, (list, tuple)):
        raise ValueError('pairings must be a list or tuple of table groups')

    table_count = len(pairings)
    if table_count < 1 or table_count > 4:
        raise ValueError('pairings must contain between 1 and 4 tables')

    reserve_substitutions = _get_reserve_substitutions(reserve1, reserve2, reserve3)

    default_output_dir = template_path.parent / 'Scorebladen'
    output_dir = Path(output_dir) if output_dir else default_output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    output_filename = f"scorebladen{play_date.strftime('%Y%m%d')}.xlsx"
    output_path = output_dir / output_filename
    shutil.copy(template_path, output_path)

    wb = load_workbook(output_path)

    for table_index in range(4):
        sheet_name = f'Tafel {table_index + 1}'
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Expected sheet '{sheet_name}' in template")
        ws = wb[sheet_name]
        _clear_scoreblad_sheet(ws)

        if table_index >= table_count:
            continue

        group = pairings[table_index]
        if not isinstance(group, (list, tuple)) or len(group) != 4:
            raise ValueError('Each table grouping must be a list or tuple of 4 player names')
        players = [str(player) for player in group]

        for cell in SCOREBLAD_DATE_CELLS:
            ws[cell].value = play_date

        # Write the three unique 2v2 matchups
        matchups = [
            (players[0], players[1], players[2], players[3]),
            (players[0], players[2], players[1], players[3]),
            (players[0], players[3], players[1], players[2]),
        ]

        for block, matchup in zip(SCOREBLAD_NAME_BLOCKS, matchups):
            for cell, player in zip(block, matchup):
                cell_value = reserve_substitutions.get(player, player)
                ws[cell].value = cell_value
                if player.casefold().startswith('reserve '):
                    ws[cell].fill = copy(RESERVE_FILL)
                    font = copy(ws[cell].font)
                    font.color = RESERVE_TEXT_COLOR
                    ws[cell].font = font

    wb.save(output_path)

    print(f"Saved filled scorebladen to {output_path}")
    return str(output_path)


def score_group(group, pair_counts):
    """
    Calculate score for a single group using MAX scoring.
    Score = the highest pair count in the group (worst pair).
    Lower score is better - we want to avoid high-count pairs.
    
    Example: If group has pairs with counts [0, 0, 1, 0, 5, 0],
    the score is 5 (the worst pairing in this group).
    """
    pair_scores = [pair_counts.get(tuple(sorted(pair)), 0) 
                   for pair in it.combinations(group, 2)]
    
    # Return the maximum pair count (worst pair) in this group
    # If all pairs are new (count=0), score will be 0 (best possible)
    return max(pair_scores) if pair_scores else 0

def make_groups_greedy(pool, group_size, csv_filename, max_configs=20):
    """
    Generate table configurations greedily, prioritizing unused/low-count pairs.
    
    SCORING METHOD:
    - Each table's score = MAX pair count in that table (worst pairing)
    - Config score = MAX across all tables (worst pair in entire config)
    - Lower score is better - minimizes the worst pairing overall
    
    Returns top configurations sorted by max pair count (lowest first).
    """
    if len(pool) % group_size != 0:
        raise ValueError("Aantal deelnemers is niet deelbaar door de groep grootte")
    
    pair_counts = load_pair_counts(csv_filename)
    configs = []
    
    # Try multiple starting points for diversity
    for seed in range(max_configs * 3):  # Try more to get best 20
        remaining = pool[:]
        config = []
        used_pairs = set()
        
        # Shuffle for different starting points
        import random
        random.seed(seed)
        random.shuffle(remaining)
        
        while remaining:
            best_group = None
            best_key = None
            
            # Try combinations prioritizing low pair counts
            attempts = 0
            for group in it.combinations(remaining, group_size):
                attempts += 1
                if attempts > 500:  # Limit search
                    break
                
                # Check Reserve constraint
                reserve_count = sum(1 for name in group if "Reserve" in name)
                if reserve_count > 1:
                    continue
                
                group_pairs = tuple(it.combinations(group, 2))
                group_score = score_group(group, pair_counts)
                group_sum = sum(pair_counts.get(tuple(sorted(pair)), 0) for pair in group_pairs)
                group_repeat_count = sum(
                    1 for pair in group_pairs
                    if pair_counts.get(tuple(sorted(pair)), 0) > 0
                )
                
                # Penalize reusing pairs within this config
                pairs_in_group = set(tuple(sorted(pair)) for pair in group_pairs)
                overlap = len(pairs_in_group & used_pairs)
                
                # Primary objective: minimize worst pair count.
                # Secondary objective: minimize repeated pairs and total pair counts.
                key = (group_score, overlap, group_repeat_count, group_sum)
                if best_key is None or key < best_key:
                    best_key = key
                    best_group = group
            
            if best_group is None:
                break  # No valid group found
            
            config.append(best_group)
            used_pairs.update(tuple(sorted(pair)) 
                            for pair in it.combinations(best_group, 2))
            remaining = [x for x in remaining if x not in best_group]
        
        if len(config) == len(pool) // group_size:  # Complete config
            # Sort players alphabetically within each group
            sorted_config = [tuple(sorted(g)) for g in config]
            
            # Calculate overall MAX score (worst pair in entire config)
            # Also compute secondary tie-breakers for equivalent max scores.
            max_pair_count = 0
            total_pair_sum = 0
            total_repeat_count = 0
            for group in sorted_config:
                group_pairs = [pair for pair in it.combinations(group, 2)
                              if not any("Reserve" in p for p in pair)]
                if group_pairs:
                    group_max = max(pair_counts.get(tuple(sorted(pair)), 0)
                                   for pair in group_pairs)
                    max_pair_count = max(max_pair_count, group_max)
                    total_pair_sum += sum(pair_counts.get(tuple(sorted(pair)), 0)
                                          for pair in group_pairs)
                    total_repeat_count += sum(
                        1 for pair in group_pairs
                        if pair_counts.get(tuple(sorted(pair)), 0) > 0
                    )
            
            # Score = worst pair count in entire configuration
            configs.append((max_pair_count, total_repeat_count, total_pair_sum, sorted_config))
    
    # Remove duplicates and sort by maximum pair count (lowest first), then by repeat count and sum.
    unique_configs = []
    seen = set()
    for score, repeat_count, total_sum, config in sorted(configs, key=lambda x: (x[0], x[1], x[2])):
        # Create hashable representation
        config_key = tuple(sorted(tuple(sorted(g)) for g in config))
        if config_key not in seen:
            seen.add(config_key)
            unique_configs.append((score, config))
            if len(unique_configs) >= max_configs:
                break
    
    return unique_configs[:max_configs]


# Alternative: Pure greedy (fastest, single best solution)
def make_groups_simple_greedy(pool, group_size, csv_filename):
    """
    Single-pass greedy: always pick group with lowest MAX pair count.
    
    SCORING METHOD:
    - Each table's score = MAX pair count in that table (worst pairing)
    - Config score = MAX across all tables (worst pair in entire config)
    - Lower score is better - minimizes the worst pairing overall
    
    Fastest option, returns one good configuration.
    """
    if len(pool) % group_size != 0:
        raise ValueError("Aantal deelnemers is niet deelbaar door de groep grootte")
    
    pair_counts = load_pair_counts(csv_filename)
    remaining = pool[:]
    config = []
    
    while remaining:
        best_group = None
        best_score = float('inf')
        
        for group in it.combinations(remaining, group_size):
            reserve_count = sum(1 for name in group if "Reserve" in name)
            if reserve_count > 1:
                continue
            
            # Score = MAX pair count in this group (worst pairing)
            score = score_group(group, pair_counts)
            if score < best_score:
                best_score = score
                best_group = group
        
        if best_group is None:
            raise ValueError("Kan geen geldige groep vinden")
        
        config.append(best_group)
        remaining = [x for x in remaining if x not in best_group]
    
    # Sort players alphabetically within each group
    sorted_config = [tuple(sorted(g)) for g in config]
    
    # Calculate overall MAX score (worst pair in entire config)
    # Only consider pairs without Reserve players
    max_pair_count = 0
    for group in sorted_config:
        # Only count pairs without Reserve players
        group_pairs = [pair for pair in it.combinations(group, 2)
                      if not any("Reserve" in p for p in pair)]
        if group_pairs:
            # Max pair count in this group
            group_max = max(pair_counts.get(tuple(sorted(pair)), 0) 
                           for pair in group_pairs)
            # Track the overall maximum across all groups
            max_pair_count = max(max_pair_count, group_max)
    
    # Score = worst pair count in entire configuration
    return [(max_pair_count, sorted_config)]


def update_pairs_in_csv(config, csv_filename, history_filename=None, play_date=None):
    """
    Update CSV with pairs from the chosen configuration.
    Each group of 4 players generates 6 pairs.
    Skips pairs that include Reserve players.
    Optionally writes the chosen configuration to a history file.
    """
    # Load existing counts
    pair_counts = load_pair_counts(csv_filename)
    
    # Add pairs from chosen configuration (excluding Reserve players)
    pairs_added = 0
    for group in config:
        for pair in it.combinations(sorted(group), 2):
            # Skip pairs with Reserve players
            if any("Reserve" in player for player in pair):
                continue
            
            pair_key = tuple(sorted(pair))
            pair_counts[pair_key] += 1
            pairs_added += 1
    
    # Write back to CSV
    with open(csv_filename, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Player1', 'Player2', 'Count'])  # Header
        for (p1, p2), count in sorted(pair_counts.items()):
            writer.writerow([p1, p2, count])

    if history_filename:
        save_pairings_history(config, play_date=play_date, history_filename=history_filename)

    print(f"Updated {csv_filename} with {pairs_added} pairs (Reserve pairs excluded)")


def delete_pairings_by_date(play_date, csv_filename=str(DATA_DIR / 'pairings.csv'), history_filename=str(DATA_DIR / 'pairings_history.csv')):
    """Delete pairings for a given play date from history and decrement counts."""
    history_rows = []
    removed_rows = []

    try:
        with open(history_filename, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header is None:
                print(f"No history found in {history_filename}.")
                return []
            for row in reader:
                if len(row) < 3:
                    continue
                row_date = row[0].strip()
                if row_date == play_date:
                    removed_rows.append([row_date, row[1].strip(), row[2].strip()])
                else:
                    history_rows.append(row)
    except FileNotFoundError:
        print(f"Warning: History file '{history_filename}' not found.")
        return []

    if not removed_rows:
        print(f"No pairings found for date {play_date} in {history_filename}.")
        return []

    # Load current pair counts and decrement for removed tables
    pair_counts = load_pair_counts(csv_filename)
    for _, _, players in removed_rows:
        participants = [p.strip() for p in players.split('|') if p.strip()]
        for pair in it.combinations(participants, 2):
            # Skip pairs with Reserve players, same as when adding pairs
            if any('Reserve' in player for player in pair):
                continue
            pair_key = tuple(sorted(pair))
            current_count = pair_counts.get(pair_key, 0)
            pair_counts[pair_key] = max(0, current_count - 1)

    # Write updated pair counts back to CSV
    with open(csv_filename, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Player1', 'Player2', 'Count'])
        for (p1, p2), count in sorted(pair_counts.items()):
            writer.writerow([p1, p2, count])

    # Write updated history file without removed date entries
    with open(history_filename, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(history_rows)

    print(f"Deleted pairings for date {play_date} from {history_filename} and updated {csv_filename}.")
    return removed_rows